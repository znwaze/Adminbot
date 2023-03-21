from functools import wraps
import telebot
import openpyxl
import os
from telegram import KeyboardButton
from datetime import datetime
from telebot.types import Message

# инициализация бота
bot = telebot.TeleBot('5730517992:AAFl-D-KSaPciyc_WMb4uUquci4mVnEYKx4')


# обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    # создание кнопок
    keyboard = telebot.types.ReplyKeyboardMarkup(True, False)
    ventilation_button = telebot.types.KeyboardButton('Вентиляция')
    electrical_button = telebot.types.KeyboardButton('Электрика')
    plumbing_button = telebot.types.KeyboardButton('Сантехника')
    status_button = KeyboardButton('Получить статус заявки')
    download_button = KeyboardButton('Скачать файл')
    answer_button = KeyboardButton('Ответить на заявку')
    keyboard.add(ventilation_button, electrical_button, plumbing_button,status_button,download_button,answer_button)
    
    # отправка сообщения с кнопками
    bot.send_message(message.chat.id, 'Выберите категорию проблемы:', reply_markup=keyboard)
# обработчик выбора категории проблемы
@bot.message_handler(func=lambda message: message.text in ['Вентиляция', 'Электрика', 'Сантехника'])
def problem_category(message):
    # сохранение категории проблемы в переменную
    category = message.text
    
    # запрос данных о пользователе
    bot.send_message(message.chat.id, 'Введите фамилию:')
    bot.register_next_step_handler(message, lambda m: user_last_name(m, category))

# обработчик ввода фамилии
def user_last_name(message, category):
    # сохранение фамилии пользователя в переменную
    last_name = message.text
    
    # запрос имени
    bot.send_message(message.chat.id, 'Введите имя:')
    bot.register_next_step_handler(message, lambda m: user_first_name(m, last_name, category))

# обработчик ввода имени
def user_first_name(message, last_name, category):
    # сохранение имени пользователя в переменную
    first_name = message.text
    
    # запрос отчества
    bot.send_message(message.chat.id, 'Введите отчество:')
    bot.register_next_step_handler(message, lambda m: user_middle_name(m, last_name, first_name, category))

# обработчик ввода отчества
def user_middle_name(message, last_name, first_name, category):
    # сохранение отчества пользователя в переменную
    middle_name = message.text
    
    # запрос информации о проблеме
    bot.send_message(message.chat.id, 'Введите информацию о проблеме:')
    bot.register_next_step_handler(message, lambda m: problem_info(m, last_name, first_name, middle_name, category))


# обработчик ввода информации о проблеме
def problem_info(message, last_name, first_name, middle_name, category):
    # сохранение информации о проблеме в переменную
    info = message.text
    
    # запрос кабинета
    bot.send_message(message.chat.id, 'Введите номер кабинета:')
    bot.register_next_step_handler(message, lambda m: room_number(m, last_name, first_name, middle_name, category, info))

#обработчик ввода номера кабинета

def room_number(message, last_name, first_name, middle_name, category, info):
    # сохранение номера кабинета в переменную
    room = message.text
    # сохранение даты и времени создания заявки

    date_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    # создание новой строки в таблице Excel
    wb = openpyxl.load_workbook('E:\\СИИИИИИИИИ\\example.xlsx')
    sheet = wb.active
    max_row = sheet.max_row
    sheet.cell(row=max_row+1, column=1, value=max_row)  # ID заявки равен номеру строки
    sheet.cell(row=max_row+1, column=2, value=last_name)
    sheet.cell(row=max_row+1, column=3, value=first_name)
    sheet.cell(row=max_row+1, column=4, value=middle_name)
    sheet.cell(row=max_row+1, column=5, value=category)
    sheet.cell(row=max_row+1, column=6, value=info)
    sheet.cell(row=max_row+1, column=7, value=room)
    sheet.cell(row=max_row+1, column=8, value=date_time)

    # отправка сообщения с информацией и кабинетом в чат, выбранный пользователем
    chat_id = None
    if category == 'Вентиляция':
        chat_id = '-1001664744852'
    elif category == 'Электрика':
        chat_id = '-1001837707549'
    elif category == 'Сантехника':
        chat_id = '-1001800170130'
    bot.send_message(chat_id, f'{last_name} {first_name} {middle_name} {info} Кабинет {room}\nID заявки: {max_row}')

    wb.save('E:\\СИИИИИИИИИ\\example.xlsx')

    # отправка сообщения о создании заявки
    bot.send_message(message.chat.id, f'Ваша заявка успешно создана!\nID заявки: {max_row}')


# обработчик выбора заявки для ответа
@bot.message_handler(func=lambda message: message.text == 'Ответить на заявку')
def reply_to_request(message):
    # запрос ID заявки
    bot.send_message(message.chat.id, 'Введите ID заявки, на которую хотите ответить:')
    bot.register_next_step_handler(message, lambda m: request_id(m))

# обработчик ввода ID заявки
def request_id(message):
    # сохранение ID заявки в переменную
    request_id = message.text
    
    # запрос ответа
    bot.send_message(message.chat.id, 'Введите ответ на заявку:')
    bot.register_next_step_handler(message, lambda m: reply_text(m, request_id))

# обработчик ввода текста ответа
def reply_text(message, request_id):
    # сохранение текста ответа в переменную
    reply_text = message.text
    
    # добавление ответа в таблицу Excel
    add_reply_to_excel(request_id, reply_text)
    
    # отправка подтверждения о сохранении ответа
    bot.send_message(message.chat.id, f'Ответ на заявку {request_id} успешно сохранен в таблице')

def add_reply_to_excel(request_id, reply_text):
    # загрузка таблицы Excel
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.active
    
    # поиск строки с нужным ID заявки
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == int(request_id):
            # добавление ответа в ячейку
            row[-1].value = reply_text
            break
    
    # сохранение изменений в Excel-файле
    wb.save('E:\\СИИИИИИИИИ\\example.xlsx')
    wb.close()



# Функция обработки сообщения с ID заявки
@bot.message_handler(func=lambda message: message.text.isdigit())
def handle_id(message: Message):
    # чтение таблицы Excel
    wb = openpyxl.load_workbook('E:\\СИИИИИИИИИ\\example.xlsx')
    sheet = wb.active

    # поиск заявки по ID
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == int(message.text):
            # отправка информации о заявке в чат
            bot.send_message(message.chat.id, f'ID заявки: {row[0]}\nФИО: {row[1]} {row[2]} {row[3]}\nКатегория: {row[4]}\nОписание: {row[5]}\nКабинет: {row[6]}\nСтатус: {row[8]}')
            break
    else:
        # если заявка не найдена, отправляем сообщение об ошибке
        bot.send_message(message.chat.id, 'Заявка с таким ID не найдена')
    
    wb.close()

    # возврат к выбору категории проблемы
    keyboard = telebot.types.ReplyKeyboardMarkup(True, False)
    ventilation_button = telebot.types.KeyboardButton('Вентиляция')
    electrical_button = telebot.types.KeyboardButton('Электрика')
    plumbing_button = telebot.types.KeyboardButton('Сантехника')
    status_button = KeyboardButton('Получить статус заявки')
    download_button = KeyboardButton('Скачать файл')
    answer_button = KeyboardButton('Ответить на заявку')
    keyboard.add(ventilation_button, electrical_button, plumbing_button,status_button,download_button,answer_button)
    
    bot.send_message(message.chat.id, 'Выберите категорию проблемы:', reply_markup=keyboard)


def restricted_access(func):
    @wraps(func)
    def wrapped(message, *args, **kwargs):
        allowed_users = [448651813, 1547587786]  # Список пользователей, которые могут скачивать файл
        if message.from_user.id not in allowed_users:
            bot.send_message(message.chat.id, "Извините, вы не можете скачивать файлы.")
            return
        return func(message, *args, **kwargs)
    return wrapped

# Обработчик нажатия на кнопку "Скачать файл"
@bot.message_handler(func=lambda message: message.text == 'Скачать файл')
@restricted_access
def handle_download(message):
    # Проверяем, что файл существует
    if os.path.isfile("E:\\СИИИИИИИИИ\\example.xlsx"):
        # Отправляем файл пользователю
        with open("E:\\СИИИИИИИИИ\\example.xlsx", 'rb') as f:
            bot.send_document(message.chat.id, f)
    else:
        # Отправляем сообщение об ошибке, если файл не найден
        bot.send_message(message.chat.id, "Файл не найден")

#обработчик команды /help
@bot.message_handler(commands=['help'])
def help(message):
    # отправка сообщения со списком доступных команд
    bot.send_message(message.chat.id, '/start - начать работу с ботом\n/help - список доступных команд\nБот работае за счет кнопок снизу, что бы заполнить заявку нажмите на кнопку снизу, это Вентиляция, Сантехника, Электрика.')


bot.polling(none_stop=True)
